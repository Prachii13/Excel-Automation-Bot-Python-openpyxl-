import openpyxl
import os
from openpyxl.styles import Font, PatternFill

INPUT_FOLDER = 'input_files'
OUTPUT_FOLDER = 'output_files'
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def is_row_empty(row):
    return all(cell.value in (None, "") for cell in row)

def clean_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Remove empty rows
    for row in reversed(range(1, sheet.max_row + 1)):
        if is_row_empty(sheet[row]):
            sheet.delete_rows(row)

    # Add Total column if numeric data is found
    last_col = sheet.max_column + 1
    sheet.cell(row=1, column=last_col).value = "Total"
    for i in range(2, sheet.max_row + 1):
        row_sum = 0
        for j in range(2, sheet.max_column):  # Skip name/id column
            val = sheet.cell(i, j).value
            if isinstance(val, (int, float)):
                row_sum += val
        sheet.cell(i, last_col).value = row_sum

    # Highlight totals > threshold
    highlight = PatternFill(start_color="FF9999", fill_type="solid")
    for i in range(2, sheet.max_row + 1):
        total = sheet.cell(i, last_col).value
        if total > 100:
            sheet.cell(i, last_col).fill = highlight

    return wb

def process_all_files():
    for file in os.listdir(INPUT_FOLDER):
        if file.endswith('.xlsx'):
            path = os.path.join(INPUT_FOLDER, file)
            wb = clean_excel(path)
            output_path = os.path.join(OUTPUT_FOLDER, file)
            wb.save(output_path)
            print(f"✅ Processed: {file}")

if __name__ == "__main__":
    process_all_files()
